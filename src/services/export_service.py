# src/services/export_service.py
"""Export Service - generates Excel/CSV files and delivers via WhatsApp"""

import os
import csv
import logging
import boto3
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from services.database import Database
from services.whatsapp_client import WhatsAppClient

logger = logging.getLogger()
logger.setLevel(logging.INFO)

BUCKET_NAME = os.environ.get('GENERATED_FILES_BUCKET', 'kashia-generated-files-dev')


class ExportService:
    """Generates Excel/CSV exports and delivers them via WhatsApp"""

    def __init__(self, database=None):
        self.db = database or Database()
        self.s3 = boto3.client('s3')
        self.whatsapp = WhatsAppClient()

    def generate_monthly_excel(self, phone_number, period="month"):
        """
        Generate a professional monthly Excel report with 3 sheets.
        Returns: (filepath, filename) or None
        """
        now = datetime.now()

        if period == "month":
            start_date = now.strftime('%Y-%m-01')
            end_date = now.strftime('%Y-%m-%d')
            period_label = now.strftime('%B_%Y')
        elif period == "week":
            monday = now - timedelta(days=now.weekday())
            start_date = monday.strftime('%Y-%m-%d')
            end_date = now.strftime('%Y-%m-%d')
            period_label = f"Week_{start_date}"
        else:
            start_date = now.strftime('%Y-%m-01')
            end_date = now.strftime('%Y-%m-%d')
            period_label = now.strftime('%B_%Y')

        transactions = self.db.get_transactions_by_period(phone_number, start_date, end_date)

        if not transactions:
            return None

        wb = Workbook()

        # ---- SHEET 1: Transactions ----
        ws1 = wb.active
        ws1.title = "Transactions"

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        money_format = '#,##0'
        income_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
        expense_fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")

        headers = ['Date', 'Description', 'Type', 'Amount (NGN)', 'Category', 'Vendor']
        for col, header in enumerate(headers, 1):
            cell = ws1.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border

        for row_idx, tx in enumerate(sorted(transactions, key=lambda x: x.get('date', '')), 2):
            ws1.cell(row=row_idx, column=1, value=tx.get('date', ''))
            ws1.cell(row=row_idx, column=2, value=tx.get('description', ''))
            ws1.cell(row=row_idx, column=3, value=tx.get('type', '').title())

            amount_cell = ws1.cell(row=row_idx, column=4, value=int(tx.get('amount', 0)))
            amount_cell.number_format = money_format

            ws1.cell(row=row_idx, column=5, value=tx.get('category', ''))
            ws1.cell(row=row_idx, column=6, value=tx.get('vendor', ''))

            fill = income_fill if tx.get('type') == 'income' else expense_fill
            for col in range(1, 7):
                ws1.cell(row=row_idx, column=col).fill = fill
                ws1.cell(row=row_idx, column=col).border = border

        for col in range(1, 7):
            max_length = max(
                len(str(ws1.cell(row=r, column=col).value or ''))
                for r in range(1, ws1.max_row + 1)
            )
            ws1.column_dimensions[get_column_letter(col)].width = min(max_length + 4, 35)

        # ---- SHEET 2: Summary ----
        ws2 = wb.create_sheet("Summary")

        income = sum(int(tx.get('amount', 0)) for tx in transactions if tx.get('type') == 'income')
        expenses = sum(int(tx.get('amount', 0)) for tx in transactions if tx.get('type') == 'expense')
        profit = income - expenses

        ws2.cell(row=1, column=1, value="Financial Summary").font = Font(bold=True, size=14)
        ws2.cell(row=2, column=1, value=f"Period: {start_date} to {end_date}")

        summary_data = [
            ('Total Income', income),
            ('Total Expenses', expenses),
            ('Net Profit/Loss', profit),
            ('Total Transactions', len(transactions)),
        ]

        ws2.cell(row=4, column=1, value="Metric").font = Font(bold=True)
        ws2.cell(row=4, column=2, value="Amount (NGN)").font = Font(bold=True)

        for i, (label, value) in enumerate(summary_data, 5):
            ws2.cell(row=i, column=1, value=label)
            cell = ws2.cell(row=i, column=2, value=value)
            if isinstance(value, int):
                cell.number_format = money_format

        ws2.column_dimensions['A'].width = 25
        ws2.column_dimensions['B'].width = 20

        # ---- SHEET 3: Category Breakdown ----
        ws3 = wb.create_sheet("Category Breakdown")

        categories = {}
        for tx in transactions:
            if tx.get('type') == 'expense':
                cat = tx.get('category', 'Other')
                categories[cat] = categories.get(cat, 0) + int(tx.get('amount', 0))

        sorted_cats = sorted(categories.items(), key=lambda x: x[1], reverse=True)

        ws3.cell(row=1, column=1, value="Expense Breakdown by Category").font = Font(bold=True, size=14)

        cat_headers = ['Category', 'Amount (NGN)', 'Percentage', 'Transactions']
        for col, header in enumerate(cat_headers, 1):
            cell = ws3.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill

        total_expense = sum(categories.values())
        for i, (cat, amount) in enumerate(sorted_cats, 4):
            ws3.cell(row=i, column=1, value=cat)
            ws3.cell(row=i, column=2, value=amount).number_format = money_format
            pct = (amount / total_expense * 100) if total_expense > 0 else 0
            ws3.cell(row=i, column=3, value=f"{pct:.1f}%")
            count = len([tx for tx in transactions if tx.get('category') == cat and tx.get('type') == 'expense'])
            ws3.cell(row=i, column=4, value=count)

        ws3.column_dimensions['A'].width = 25
        ws3.column_dimensions['B'].width = 18
        ws3.column_dimensions['C'].width = 12
        ws3.column_dimensions['D'].width = 14

        filename = f"Kashia_Report_{period_label}.xlsx"
        filepath = f"/tmp/{filename}"
        wb.save(filepath)

        logger.info(f"Excel report generated: {filepath}")
        return filepath, filename

    def generate_csv(self, phone_number, period="all"):
        """Generate a CSV export of all transactions."""
        if period == "all":
            transactions = self.db.get_transactions(phone_number, limit=1000)
        else:
            now = datetime.now()
            start_date = now.strftime('%Y-%m-01')
            end_date = now.strftime('%Y-%m-%d')
            transactions = self.db.get_transactions_by_period(phone_number, start_date, end_date)

        if not transactions:
            return None, None

        filename = f"Kashia_Transactions_{datetime.now().strftime('%Y%m%d')}.csv"
        filepath = f"/tmp/{filename}"

        with open(filepath, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)
            writer.writerow(['Date', 'Description', 'Type', 'Amount', 'Category', 'Sub-Category', 'Vendor'])

            for tx in sorted(transactions, key=lambda x: x.get('date', '')):
                writer.writerow([
                    tx.get('date', ''),
                    tx.get('description', ''),
                    tx.get('type', ''),
                    tx.get('amount', 0),
                    tx.get('category', ''),
                    tx.get('sub_category', ''),
                    tx.get('vendor', ''),
                ])

        logger.info(f"CSV generated: {filepath}")
        return filepath, filename

    def generate_contacts_export(self, phone_number):
        """Export all contacts as Excel."""
        contacts = self.db.get_contacts(phone_number)

        if not contacts:
            return None, None

        wb = Workbook()
        ws = wb.active
        ws.title = "Contacts"

        headers = ['Name', 'Type', 'Total Paid (NGN)', 'Total Received (NGN)', 'Transactions', 'Last Transaction']
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill

        for i, contact in enumerate(contacts, 2):
            ws.cell(row=i, column=1, value=contact.get('name', ''))
            ws.cell(row=i, column=2, value=contact.get('type', '').title())
            ws.cell(row=i, column=3, value=int(contact.get('total_paid', 0)))
            ws.cell(row=i, column=4, value=int(contact.get('total_received', 0)))
            ws.cell(row=i, column=5, value=int(contact.get('transaction_count', 0)))
            ws.cell(row=i, column=6, value=contact.get('last_transaction_date', ''))

        for col in range(1, 7):
            ws.column_dimensions[get_column_letter(col)].width = 18

        filename = f"Kashia_Contacts_{datetime.now().strftime('%Y%m%d')}.xlsx"
        filepath = f"/tmp/{filename}"
        wb.save(filepath)

        return filepath, filename

    def upload_to_s3(self, filepath, filename):
        """Upload a file to S3 and return a presigned URL (24hr expiry)."""
        try:
            s3_key = f"exports/{datetime.now().strftime('%Y%m%d')}/{filename}"
            self.s3.upload_file(filepath, BUCKET_NAME, s3_key)

            presigned_url = self.s3.generate_presigned_url(
                'get_object',
                Params={'Bucket': BUCKET_NAME, 'Key': s3_key},
                ExpiresIn=86400
            )

            logger.info(f"Uploaded to S3: {s3_key}")
            return presigned_url

        except Exception as e:
            logger.error(f"S3 upload error: {e}")
            return None

    def deliver_file(self, phone_number, filepath, filename, caption=""):
        """Full pipeline: upload to S3 then send to user via WhatsApp."""
        url = self.upload_to_s3(filepath, filename)
        if not url:
            return False, None

        success = self.whatsapp.send_document(phone_number, url, filename, caption)

        try:
            os.remove(filepath)
        except:
            pass

        return success, url

    def handle_export_request(self, phone_number, export_type):
        """
        Handle an export request end-to-end.
        Args: export_type: "month", "csv", "contacts"
        Returns: list of response dicts
        """
        if export_type in ["month", "export_month", "1"]:
            result = self.generate_monthly_excel(phone_number)
            if result:
                filepath, filename = result
                self.deliver_file(phone_number, filepath, filename,
                                  caption="Your monthly report - open in Excel or Google Sheets!")
                return [{"type": "text", "content": "File sent! Check your chat for the Excel file."}]
            else:
                return [{"type": "text", "content": "No transactions this month yet."}]

        elif export_type in ["csv", "export_csv", "2", "full"]:
            filepath, filename = self.generate_csv(phone_number)
            if filepath:
                self.deliver_file(phone_number, filepath, filename,
                                  caption="Full transaction history (CSV)")
                return [{"type": "text", "content": "CSV file sent!"}]
            else:
                return [{"type": "text", "content": "No transactions to export."}]

        elif export_type in ["contacts", "export_contacts", "3"]:
            filepath, filename = self.generate_contacts_export(phone_number)
            if filepath:
                self.deliver_file(phone_number, filepath, filename,
                                  caption="Your contacts list")
                return [{"type": "text", "content": "Contacts list sent!"}]
            else:
                return [{"type": "text", "content": "No contacts to export yet."}]

        else:
            return [{"type": "text", "content": "Unknown export type. Try: export month, export csv, or export contacts"}]
    def handle_filtered_export(self, phone_number, filter_type, start_date, end_date, period_label, fmt='excel'):
        """Export filtered transaction data as Excel or PDF"""
        try:
            transactions = self.db.get_transactions_by_period(phone_number, start_date, end_date)

            # Apply same filters as the report
            COGS_CATEGORIES = ['Goods & Stock', 'Production & Manufacturing', 'Service Costs']

            if filter_type == 'my_sales':
                filtered = [tx for tx in transactions
                           if tx.get('type') == 'income'
                           and 'Debt payment' not in tx.get('description', '')]
                label = 'Sales'
            elif filter_type == 'my_purchases':
                filtered = [tx for tx in transactions
                           if tx.get('type') == 'expense'
                           and tx.get('category', '') in COGS_CATEGORIES]
                label = 'Purchases'
            elif filter_type == 'my_expenses':
                filtered = [tx for tx in transactions
                           if tx.get('type') == 'expense'
                           and tx.get('category', '') not in COGS_CATEGORIES]
                label = 'Expenses'
            else:
                filtered = transactions
                label = 'Transactions'

            if not filtered:
                return [{"type": "text", "content": f"No {label.lower()} to export for {period_label}."}]

            user = self.db.get_user(phone_number)
            business_name = user.get('business_name', 'Business') if user else 'Business'

            if fmt == 'excel':
                return self._export_filtered_excel(phone_number, filtered, label, period_label, business_name)
            else:
                return self._export_filtered_pdf(phone_number, filtered, label, period_label, business_name)

        except Exception as e:
            import logging
            logging.getLogger(__name__).error(f"Filtered export error: {e}")
            return [{"type": "text", "content": "Sorry, export failed. Please try again."}]

    def _export_filtered_excel(self, phone_number, transactions, label, period_label, business_name):
        """Generate Excel file for filtered transactions"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from datetime import datetime
        import os

        wb = Workbook()
        ws = wb.active
        ws.title = f"{label} - {period_label}"

        # Header
        ws['A1'] = business_name
        ws['A1'].font = Font(bold=True, size=14)
        ws['A2'] = f"{label} Report - {period_label}"
        ws['A2'].font = Font(size=11)
        ws['A3'] = f"Generated: {datetime.now().strftime('%d %B %Y %H:%M')}"

        # Column headers
        headers = ['Date', 'Description', 'Category', 'Customer/Vendor', 'Amount (NGN)']
        header_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill

        # Data rows
        total = 0
        for row_idx, tx in enumerate(sorted(transactions, key=lambda x: x.get('date', ''), reverse=True), 6):
            amount = int(tx.get('amount', 0))
            total += amount
            ws.cell(row=row_idx, column=1, value=tx.get('date', ''))
            ws.cell(row=row_idx, column=2, value=tx.get('description', '')[:50])
            ws.cell(row=row_idx, column=3, value=tx.get('category', ''))
            ws.cell(row=row_idx, column=4, value=tx.get('vendor', ''))
            ws.cell(row=row_idx, column=5, value=amount)

        # Total row
        total_row = 6 + len(transactions)
        ws.cell(row=total_row, column=4, value='TOTAL').font = Font(bold=True)
        ws.cell(row=total_row, column=5, value=total).font = Font(bold=True)

        # Column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 15

        # Save
        clean_period = period_label.replace(' ', '_').replace('/', '-')
        filename = f"{label}_{clean_period}.xlsx"
        filepath = f"/tmp/{filename}"
        wb.save(filepath)

        # Deliver
        self.deliver_file(phone_number, filepath, filename,
                         caption=f"{label} - {period_label} ({len(transactions)} transactions)")
        return [{"type": "text", "content": f"\u2705 Excel exported!\n\n{label} \u2014 {period_label}\n{len(transactions)} transactions | Total: NGN {total:,}\n\n\U0001f4ce Check your chat for the file."}]

    def _export_filtered_pdf(self, phone_number, transactions, label, period_label, business_name):
        """Generate PDF report for filtered transactions"""
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import cm, mm
        from reportlab.lib import colors
        from reportlab.lib.colors import HexColor, black, white
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet
        from datetime import datetime

        filename = f"{label}_{period_label.replace(' ', '_').replace('/', '-')}.pdf"
        filepath = f"/tmp/{filename}"

        doc = SimpleDocTemplate(filepath, pagesize=A4,
                               rightMargin=2*cm, leftMargin=2*cm,
                               topMargin=2*cm, bottomMargin=2*cm)

        styles = getSampleStyleSheet()
        story = []

        # Header
        story.append(Paragraph(f"<b>{business_name}</b>", styles['Title']))
        story.append(Paragraph(f"{label} Report - {period_label}", styles['Heading2']))
        story.append(Paragraph(f"Generated: {datetime.now().strftime('%d %B %Y %H:%M')}", styles['Normal']))
        story.append(Spacer(1, 10*mm))

        # Summary
        total = sum(int(tx.get('amount', 0)) for tx in transactions)
        story.append(Paragraph(f"<b>Total:</b> NGN {total:,}", styles['Normal']))
        story.append(Paragraph(f"<b>Transactions:</b> {len(transactions)}", styles['Normal']))
        story.append(Spacer(1, 8*mm))

        # Table
        table_data = [['Date', 'Description', 'Category', 'Vendor', 'Amount (NGN)']]
        for tx in sorted(transactions, key=lambda x: x.get('date', ''), reverse=True):
            amount = int(tx.get('amount', 0))
            table_data.append([
                tx.get('date', ''),
                tx.get('description', '')[:35],
                tx.get('category', ''),
                tx.get('vendor', ''),
                f"NGN {amount:,}"
            ])
        table_data.append(['', '', '', 'TOTAL', f"NGN {total:,}"])

        t = Table(table_data, colWidths=[2.5*cm, 5*cm, 4*cm, 3.5*cm, 3*cm])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), HexColor('#2c3e50')),
            ('TEXTCOLOR', (0, 0), (-1, 0), white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('ALIGN', (-1, 0), (-1, -1), 'RIGHT'),
            ('GRID', (0, 0), (-1, -2), 0.5, HexColor('#cccccc')),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('LINEABOVE', (0, -1), (-1, -1), 1.5, black),
            ('TOPPADDING', (0, 0), (-1, -1), 5),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ]))
        story.append(t)
        story.append(Spacer(1, 10*mm))
        story.append(Paragraph("Generated by Kashia - AI Bookkeeping for Nigerian Businesses", styles['Normal']))

        doc.build(story)

        self.deliver_file(phone_number, filepath, filename,
                         caption=f"{label} - {period_label}")
        return [{"type": "text", "content": f"\u2705 PDF exported!\n\n{label} \u2014 {period_label}\n{len(transactions)} transactions | Total: NGN {total:,}\n\n\U0001f4ce Check your chat for the file."}]
    def export_full_history_csv(self, phone_number):
        """Export ALL transactions as a CSV file"""
        import csv
        from datetime import datetime
        
        # Get all transactions (use a very wide date range)
        transactions = self.db.get_transactions_by_period(phone_number, '2020-01-01', '2030-12-31')
        
        if not transactions:
            return None, None
        
        user = self.db.get_user(phone_number)
        business_name = user.get('business_name', 'Business') if user else 'Business'
        
        filename = f"{business_name.replace(' ', '_')}_Full_History.csv"
        filepath = f"/tmp/{filename}"
        
        with open(filepath, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            # Header
            writer.writerow(['Date', 'Type', 'Category', 'Description', 'Vendor/Customer', 
                           'Amount', 'Quantity', 'Unit Cost', 'Brand', 'Payment'])
            
            # Sort chronologically
            transactions.sort(key=lambda x: x.get('created_at', x.get('date', '')))
            
            for tx in transactions:
                writer.writerow([
                    tx.get('date', ''),
                    tx.get('type', ''),
                    tx.get('category', ''),
                    tx.get('description', ''),
                    tx.get('vendor', ''),
                    tx.get('amount', 0),
                    tx.get('quantity', ''),
                    tx.get('unit_cost', ''),
                    tx.get('brand', ''),
                    tx.get('payment_method', 'cash'),
                ])
        
        return filepath, filename
